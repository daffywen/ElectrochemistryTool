"""
线性扫描伏安法(LSV)数据处理模块
"""
import os
import sys
import numpy as np
import logging
import re
from typing import Tuple, List, Optional, Any
from datetime import datetime

try:
    from tqdm import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False
    print("提示: 安装tqdm包可以显示进度条。可以运行: pip install tqdm")

# 检测是否在 GUI 环境中运行（如 PyInstaller 打包后的应用）
def is_gui_mode():
    """判断是否在无控制台的GUI环境中运行"""
    # 检查是否通过 PyInstaller 打包
    if getattr(sys, 'frozen', False):
        return True
    # 检查是否有可用的控制台 
    try:
        # 尝试向标准输出写入，如果失败则可能在GUI模式中
        sys.stdout.write("")
        return False
    except (AttributeError, IOError):
        return True

logger = logging.getLogger(__name__)
    
# 在GUI模式下禁用tqdm
if is_gui_mode():
    TQDM_AVAILABLE = False
    logger.info("检测到GUI模式，已禁用进度条显示")

# 导入openpyxl样式到模块级别
try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# 导入共享模块
try:
    from .common import file_utils, excel_utils
except ImportError:
    # 如果作为独立模块运行
    try:
        import sys
        sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from electrochemistry.common import file_utils, excel_utils
    except ImportError:
        print("无法导入共享模块，请确保目录结构正确")
        raise

logger = logging.getLogger(__name__)

def extract_lsv_data(filename: str) -> Tuple[List[float], List[float], Optional[str]]:
    """
    从LSV数据文件中提取电位和电流数据
    
    参数:
        filename: 文件路径
        
    返回:
        (电位列表, 电流列表, 文件标识)
    """
    try:
        with open(filename, 'r', errors='ignore') as f:
            content = f.read()
        
        # 提取文件标识符（例如"File: lsv4"中的"lsv4"）
        file_id = None
        file_match = re.search(r'File:\s*([^\r\n,]+)', content)
        if file_match:
            file_id = file_match.group(1).strip()
        else:
            # 如果找不到File标识，使用文件名作为标识
            base_name = os.path.basename(filename)
            file_id = os.path.splitext(base_name)[0]
        
        # 提取电位和电流数据
        potentials = []
        currents = []
        
        # 查找数据部分
        if 'Potential/V, Current/A' in content:
            data_section = content.split('Potential/V, Current/A')[1].strip()
        elif 'Potential' in content and 'Current' in content:
            # 尝试找到数据部分的开始
            lines = content.strip().split('\n')
            data_start = -1
            for i, line in enumerate(lines):
                if ('Potential' in line and 'Current' in line) or \
                   (re.search(r'[Pp]otential.*[Vv]', line) and re.search(r'[Cc]urrent.*[Aa]', line)):
                    data_start = i + 1
                    break
            
            if data_start > 0 and data_start < len(lines):
                data_section = '\n'.join(lines[data_start:])
            else:
                # 如果找不到明确的标题行，尝试查找数字数据段
                potential_current_pattern = re.compile(r'(-?\d+\.?\d*)\s*,\s*(-?\d+\.?\d*)')
                matches = potential_current_pattern.findall(content)
                if matches:
                    return [float(p) for p, c in matches], [float(c) for p, c in matches], file_id
                return [], [], file_id
        else:
            logger.warning(f"在文件{filename}中未找到电位和电流数据")
            return [], [], file_id
        
        lines = data_section.strip().split('\n')
        
        # 预分配数组提高性能
        potentials = []
        currents = []
        
        for line in lines:
            if line.strip():
                # 尝试多种分隔符
                if ',' in line:
                    parts = line.strip().split(',')
                elif '\t' in line:
                    parts = line.strip().split('\t')
                else:
                    parts = line.strip().split()
                
                if len(parts) >= 2:
                    try:
                        potential = float(parts[0].strip())
                        current = float(parts[1].strip())
                        potentials.append(potential)
                        currents.append(current)
                    except (ValueError, IndexError):
                        continue
        
        return potentials, currents, file_id
    
    except Exception as e:
        logger.error(f"处理文件{filename}时出错: {str(e)}")
        return [], [], None

def find_closest_values(values: List[float], targets: List[float]) -> List[int]:
    """
    找到数组中最接近目标值的元素索引
    
    参数:
        values: 要搜索的数组
        targets: 目标值列表
        
    返回:
        最接近目标值的索引列表
    """
    if not values or not targets:
        return []
        
    result_indices = []
    values_array = np.array(values)
    
    for target in targets:
        idx = np.argmin(np.abs(values_array - target))
        result_indices.append(idx)
    
    return result_indices

def process_lsv_files(file_paths: List[str], output_file: str = None, cv_data_exists: bool = False, wb=None) -> Tuple[Optional[openpyxl.Workbook], Optional[List[Any]], Optional[List[str]]]: # MODIFIED return type
    """
    处理LSV文件并准备Excel数据
    
    参数:
        file_paths: LSV文件路径列表
        output_file: 输出Excel文件路径，如果为None则只返回数据不保存
        cv_data_exists: 是否存在CV数据，影响Excel布局
        wb: 可选，现有的工作簿对象
        
    返回:
        (工作簿对象, 分析数据列表, 创建/使用的工作表名称列表)
    """
    # 提取并处理所有数据
    all_data = []
    file_ids = []
    
    # 使用进度条（如果可用）
    file_iterator = tqdm(file_paths, desc="处理LSV文件") if TQDM_AVAILABLE else file_paths
    
    for file_path in file_iterator:
        try:
            potentials, currents, file_id = extract_lsv_data(file_path)
            
            if not potentials or not currents:
                logger.warning(f"文件 {file_path} 中未找到有效数据，跳过")
                continue
                
            # 应用转换：电位+0.903，电流*-1000
            processed_potentials = [p + 0.903 for p in potentials]
            processed_currents = [c * -1000 for c in currents]
            
            all_data.append((processed_potentials, processed_currents))
            file_ids.append(file_id)
            
            logger.info(f"文件 {file_path}: 已处理 {len(potentials)} 个数据点")
            
        except Exception as e:
            logger.error(f"处理{file_path}时出错: {str(e)}")
            continue
    
    if not all_data:
        logger.error("没有找到任何可处理的LSV数据文件")
        return None, None, None # MODIFIED return
    
    # 查找最大数据点数量
    max_length = max([len(potentials) for potentials, _ in all_data]) if all_data else 0
    
    created_sheet_names = [] # ADDED: To store created/used sheet names

    # 使用提供的工作簿或创建新的
    if wb is None:
        # 为LSV数据创建新工作簿
        wb, ws, header_fill, thin_border, center_aligned, openpyxl_module = excel_utils.setup_excel_workbook("LSV Data")
        logger.info("为LSV数据创建了新的工作簿和工作表 'LSV Data'")
        created_sheet_names.append("LSV Data") # ADDED
    else:
        # 使用提供的工作簿并获取所需样式
        header_fill, thin_border, center_aligned, openpyxl_module = excel_utils.get_excel_styles()
        
        # 检查工作表是否已存在或创建新的
        worksheet_name = "LSV Data"
        if worksheet_name in wb.sheetnames:
            ws = wb[worksheet_name]
            logger.info(f"使用现有的工作表 '{worksheet_name}' 处理LSV数据")
        else:
            ws = wb.create_sheet(worksheet_name)
            logger.info(f"已创建新的工作表 '{worksheet_name}' 用于LSV数据")
        created_sheet_names.append(worksheet_name) # ADDED
    
    # 确保可以使用PatternFill和Font
    # PatternFill = openpyxl_module.styles.PatternFill # 如果使用header_fill则不需要
    # Font = openpyxl_module.styles.Font # 将使用 excel_utils.get_bold_font
    # Alignment = openpyxl_module.styles.Alignment # 将使用 excel_utils 中的 center_aligned
    bold_font = excel_utils.get_bold_font() # 使用工具函数

    current_col_lsv = 1 # 跟踪LSV数据的列
    all_analysis_data = [] # 用于存储 (file_id, op_at_10, op_at_100, op_at_200)
    
    # 为每个LSV文件创建标题和数据
    for idx, file_id in enumerate(file_ids):
        # --- 写入LSV数据部分 ---
        # 第一行: 电位和电流密度
        ws.cell(row=1, column=current_col_lsv).value = "Potential"
        ws.cell(row=1, column=current_col_lsv + 1).value = "Current Density"
        
        # 第二行: 单位
        ws.cell(row=2, column=current_col_lsv).value = "V vs. RHE"
        ws.cell(row=2, column=current_col_lsv + 1).value = "mA·cm⁻²"
        
        # 第三行: 文件标识
        ws.cell(row=3, column=current_col_lsv).value = ""
        ws.cell(row=3, column=current_col_lsv + 1).value = file_id
        
        # 应用样式到LSV表头 (rows 1-3)
        for row in range(1, 4):
            for col_offset in range(2):
                cell = ws.cell(row=row, column=current_col_lsv + col_offset)
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_aligned # 使用 excel_utils 中的 center_aligned
                # 将粗体字体应用于数据部分第1-3行中的所有表头单元格
                if bold_font:
                    cell.font = bold_font

        # 第四行: 格式化的空行 (LSV数据)
        for col_offset in range(2):
            cell = ws.cell(row=4, column=current_col_lsv + col_offset)
            cell.value = None 
            cell.fill = header_fill 
            cell.border = thin_border
            cell.alignment = center_aligned
        
        potentials, currents = all_data[idx]
        # highlight_indices = find_closest_values(currents, [10, 100]) # 用于LSV数据高亮 - 已移除

        for row_idx_offset, (potential, current) in enumerate(zip(potentials, currents)):
            actual_row = row_idx_offset + 5 # 数据从第5行开始
            potential_cell = ws.cell(row=actual_row, column=current_col_lsv, value=potential)
            current_cell = ws.cell(row=actual_row, column=current_col_lsv + 1, value=current)
            
            potential_cell.border = thin_border
            current_cell.border = thin_border
            
            # 已移除高亮逻辑
            # if row_idx_offset in highlight_indices:
            #     highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            #     potential_cell.fill = highlight_fill
            #     potential_cell.font = bold_font
            #     current_cell.fill = highlight_fill
            #     current_cell.font = bold_font
        
        # 填充LSV数据的空白单元格
        for row_idx_offset in range(len(potentials), max_length):
            actual_row = row_idx_offset + 5
            ws.cell(row=actual_row, column=current_col_lsv).value = None
            ws.cell(row=actual_row, column=current_col_lsv + 1).value = None
            ws.cell(row=actual_row, column=current_col_lsv).border = thin_border
            ws.cell(row=actual_row, column=current_col_lsv + 1).border = thin_border

        # --- 计算并存储分析数据 ---
        op_at_10_val = "N/A"
        op_at_100_val = "N/A"
        op_at_200_val = "N/A" # 新增：200 mA/cm²时的过电位
        if currents: # 确保有数据可处理
            # 查找10, 100, 和 200 mA/cm²时的电位
            closest_indices_for_analysis = find_closest_values(currents, [10.0, 100.0, 200.0])
            
            if len(closest_indices_for_analysis) >= 1: # 检查是否至少找到一个值
                try:
                    p_at_10 = potentials[closest_indices_for_analysis[0]]
                    op_at_10_val = (p_at_10 - 1.23) * 1000 # 转换为mV
                except (IndexError, TypeError): # 为安全起见添加了TypeError
                    logger.warning(f"文件 {file_id}: 无法找到10 mA·cm⁻²时的电位。")
            
            if len(closest_indices_for_analysis) >= 2:
                try:
                    p_at_100 = potentials[closest_indices_for_analysis[1]]
                    op_at_100_val = (p_at_100 - 1.23) * 1000 # 转换为mV
                except (IndexError, TypeError):
                    logger.warning(f"文件 {file_id}: 无法找到100 mA·cm⁻²时的电位。")

            if len(closest_indices_for_analysis) >= 3: # 新增：用于200 mA/cm²
                try:
                    p_at_200 = potentials[closest_indices_for_analysis[2]]
                    op_at_200_val = (p_at_200 - 1.23) * 1000 # 转换为mV
                except (IndexError, TypeError):
                    logger.warning(f"文件 {file_id}: 无法找到200 mA·cm⁻²时的电位。")
                    
        all_analysis_data.append((file_id, op_at_10_val, op_at_100_val, op_at_200_val))

        # 更新下一个LSV文件块的起始列
        current_col_lsv += 2
    
    # --- 写入集中的分析结果部分 ---
    # 在所有LSV数据之后，留出一列空隙
    analysis_section_start_col = current_col_lsv + 1

    if all_analysis_data:
        # 分析结果主表头 (Row 1)
        analysis_main_header_cell = ws.cell(row=1, column=analysis_section_start_col, value="Analysis Results") # 已改为英文
        analysis_main_header_cell.fill = header_fill
        analysis_main_header_cell.border = thin_border
        analysis_main_header_cell.font = bold_font
        analysis_main_header_cell.alignment = Alignment(horizontal='center', vertical='center')
        # 合并4列用于 "File ID", "Overpotential @10", "Overpotential @100", "Overpotential @200"
        ws.merge_cells(start_row=1, start_column=analysis_section_start_col, end_row=1, end_column=analysis_section_start_col + 3)
        # 为合并范围内的其他单元格设置样式以保持边框/填充一致
        for c_offset in range(1, 4): # 调整范围为4
            merged_cell_part = ws.cell(row=1, column=analysis_section_start_col + c_offset)
            merged_cell_part.fill = header_fill
            merged_cell_part.border = thin_border
            if bold_font: # 确保合并部分也为粗体
                merged_cell_part.font = bold_font

        # 分析结果列标题 (Row 2)
        ws.cell(row=2, column=analysis_section_start_col).value = "File ID" # 已改为英文
        ws.cell(row=2, column=analysis_section_start_col + 1).value = "Overpotential @10mA·cm⁻²" # 已改为英文
        ws.cell(row=2, column=analysis_section_start_col + 2).value = "Overpotential @100mA·cm⁻²" # 已改为英文
        ws.cell(row=2, column=analysis_section_start_col + 3).value = "Overpotential @200mA·cm⁻²" # 新列

        # 分析结果单位 (Row 3)
        ws.cell(row=3, column=analysis_section_start_col).value = ""
        ws.cell(row=3, column=analysis_section_start_col + 1).value = "(mV)"
        ws.cell(row=3, column=analysis_section_start_col + 2).value = "(mV)"
        ws.cell(row=3, column=analysis_section_start_col + 3).value = "(mV)" # 新单位

        # 样式化分析结果的表头 (Rows 2-3)
        for r in range(2, 4):
            for c_offset in range(4): # 调整范围为4
                cell = ws.cell(row=r, column=analysis_section_start_col + c_offset)
                cell.fill = header_fill
                cell.border = thin_border
                if bold_font:
                    cell.font = bold_font
                cell.alignment = center_aligned # 使用 excel_utils 中的 center_aligned

        # 第四行: 格式化的空行 (分析结果)
        for c_offset in range(4): # 调整范围为4
            cell = ws.cell(row=4, column=analysis_section_start_col + c_offset)
            cell.value = None
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_aligned

        # 写入分析数据 (从第5行开始)
        for row_idx_offset, (file_id_an, op_10, op_100, op_200) in enumerate(all_analysis_data): # 解包 op_200
            actual_row_an = row_idx_offset + 5
            
            cell_file_id_an = ws.cell(row=actual_row_an, column=analysis_section_start_col, value=file_id_an)
            cell_op_10 = ws.cell(row=actual_row_an, column=analysis_section_start_col + 1, value=op_10)
            cell_op_100 = ws.cell(row=actual_row_an, column=analysis_section_start_col + 2, value=op_100)
            cell_op_200 = ws.cell(row=actual_row_an, column=analysis_section_start_col + 3, value=op_200) # 新单元格

            # 应用边框和对齐
            for c_offset_an in range(4): # 调整范围为4
                cell_an = ws.cell(row=actual_row_an, column=analysis_section_start_col + c_offset_an)
                cell_an.border = thin_border
                if c_offset_an == 0: # 文件ID
                    # 直接创建Alignment对象
                    cell_an.alignment = openpyxl_module.styles.Alignment(horizontal='left') 
                else: # 过电位值
                    # 直接创建Alignment对象
                    cell_an.alignment = openpyxl_module.styles.Alignment(horizontal='right') 
            
            # 过电位的数字格式
            if isinstance(op_10, float):
                cell_op_10.number_format = '0.0' # mV格式，例如1位小数
            else: # 对于 "N/A"
                 cell_op_10.alignment = Alignment(horizontal='center')
            if isinstance(op_100, float):
                cell_op_100.number_format = '0.0' # mV格式
            else: # 对于 "N/A"
                 cell_op_100.alignment = Alignment(horizontal='center')
            if isinstance(op_200, float): # op_200的新格式
                cell_op_200.number_format = '0.0' 
            else: # 对于 "N/A"
                 cell_op_200.alignment = Alignment(horizontal='center')
        
        final_content_end_col = analysis_section_start_col + 3 # 调整以适应新列
    else:
        final_content_end_col = current_col_lsv -1 if current_col_lsv > 1 else 0

    # 设置列宽 (从1到最后一个有内容的列)
    # if final_content_end_col > 0:
    #     for col_idx in range(1, final_content_end_col + 1):
    #         col_letter = openpyxl_module.utils.get_column_letter(col_idx)
    #         # 检查是否为分隔列 (LSV数据之后，分析结果之前的那一列)
    #         if col_idx == current_col_lsv: # 这是分隔列
    #             ws.column_dimensions[col_letter].width = 5 # 分隔列使用较窄的宽度
    #         else:
    #             ws.column_dimensions[col_letter].width = 18 # 数据/分析列使用标准宽度
    # 使用工具函数设置列宽
    if final_content_end_col > 0:
        col_width_map = {}
        for col_idx in range(1, final_content_end_col + 1):
            col_letter = openpyxl_module.utils.get_column_letter(col_idx)
            if col_idx == current_col_lsv: # 分隔列
                col_width_map[col_letter] = 5
            else:
                col_width_map[col_letter] = 18
        excel_utils.set_column_widths(ws, col_width_map)

    logger.info("LSV数据及分析处理完成。Excel工作表已更新。")
    
    # 保存操作由主模块处理
    # if output_file and create_new_workbook_internally: 
    #     try:
    #         wb.save(output_file)
    #         logger.info(f"LSV数据已保存到 {output_file} 的 'LSV Data' 工作表")
    #     except Exception as e:
    #         logger.error(f"保存LSV数据到 {output_file} 时出错: {str(e)}")
            
    return wb, all_analysis_data, created_sheet_names # MODIFIED return

def process_all_files_from_paths(file_paths: List[str], output_file: str, folder_basename: str, wb: Optional[openpyxl.Workbook] = None) -> Tuple[Optional[openpyxl.Workbook], Optional[List[Any]], Optional[List[str]]]:
    """
    处理来自给定路径列表的所有LSV文件。
    这是从 main.py 调用的包装器。
    """
    cv_data_exists = bool(wb and "CV Data" in wb.sheetnames)

    # The process_lsv_files function in the provided snippet does not use folder_basename directly in its parameters.
    # It uses output_file, which might be related or not.
    # Assuming process_lsv_files is the main worker.
    workbook, analysis_results, sheet_names = process_lsv_files(
        file_paths=file_paths,
        output_file=output_file, 
        cv_data_exists=cv_data_exists,
        wb=wb
    )
    return workbook, analysis_results, sheet_names

def find_lsv_files(folder_path: str) -> List[str]:
    """
    在指定文件夹中查找所有可能的LSV数据文件
    通过检查文件内容判断是否为LSV文件
    """
    lsv_files = []
    
    # 获取文件夹中所有txt文件
    txt_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) 
                if f.lower().endswith('.txt')]
    
    if not txt_files:
        logger.warning(f"在{folder_path}中未找到任何.txt文件")
        return []
        
    logger.info(f"在{folder_path}中找到{len(txt_files)}个.txt文件，检查是否为LSV数据...")
    
    # 使用进度条（如果可用）
    file_iterator = tqdm(txt_files, desc="识别LSV数据文件") if TQDM_AVAILABLE else txt_files
    
    # 检查每个文件是否为LSV数据文件
    for file_path in file_iterator:
        try:
            with open(file_path, 'r', errors='ignore') as f:
                # 只读取文件头部内容用于检测（提高性能）
                header = f.read(1000)
                
                # 匹配LSV关键词
                is_lsv = ('Linear Sweep Voltammetry' in header or 
                         'LINEAR SWEEP VOLTAMMETRY' in header or
                         'LSV' in header)
                
                # 排除非LSV文件类型
                is_not_lsv = any(keyword in header for keyword in [
                    'Cyclic Voltammetry', 
                    'A.C. Impedance',
                    'CV',
                    'Chronoamperometry',
                    'Open Circuit',
                    'EIS',
                    'Tafel'
                ])
                
                # 只有确认是LSV且不包含非LSV关键词的文件才返回True
                if is_lsv and not is_not_lsv:
                    lsv_files.append(file_path)
                    logger.debug(f"文件 {file_path} 被识别为LSV数据文件")
                # elif is_lsv and is_not_lsv: # Debugging: Log files that are LSV but also contain other keywords
                #     logger.debug(f"File {file_path} contains LSV keywords but also other keywords, skipping.")
                # elif not is_lsv: # Debugging: Log files that are not LSV
                #     logger.debug(f"File {file_path} does not appear to be an LSV file.")

        except Exception as e:
            logger.error(f"检查文件 {file_path} 时出错: {str(e)}")
            continue
            
    if not lsv_files:
        logger.info(f"在 {folder_path} 中未找到LSV数据文件")
    else:
        logger.info(f"在 {folder_path} 中找到 {len(lsv_files)} 个LSV数据文件")
        
    return lsv_files

def main():
    """LSV数据处理主函数"""
    logger.info("LSV数据处理程序启动")
    
    # 提示用户选择文件夹
    folder_path = file_utils.select_folder()
    
    if not folder_path:
        logger.warning("未选择文件夹，程序退出。")
        return
    
    logger.info(f"已选择文件夹: {folder_path}")
    
    # 在选定的文件夹中查找LSV文件
    lsv_files = find_lsv_files(folder_path)
    
    if not lsv_files:
        logger.error(f"在选定的文件夹中未找到任何LSV数据文件。")
        print("\n在选定的文件夹中未找到任何LSV数据文件。请确保文件内容包含线性扫描伏安法相关信息。")
        return
    
    logger.info(f"找到 {len(lsv_files)} 个LSV数据文件:")
    for file_path in lsv_files:
        logger.info(f"  {os.path.basename(file_path)}")
    
    # 准备输出目录和文件名
    output_dir, folder_basename = file_utils.ensure_output_dir(folder_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(output_dir, f"{folder_basename}_processed_lsv_data_{timestamp}.xlsx")
    
    # 处理文件并保存结果
    try:
        processed_wb = process_all_files_from_paths(lsv_files, output_file, folder_basename, None)
        success = bool(processed_wb)
        
        if success:
            print(f"\n处理完成！结果已保存到：{output_file}")
            print(f"共处理了 {len(lsv_files)} 个LSV数据文件")
        else:
            print("\nLSV数据处理过程中发生错误，请查看日志获取详情。")
        
    except Exception as e:
        logger.error(f"发生意外错误: {str(e)}")
        print(f"\n发生意外错误: {str(e)}")
        print("\n建议解决问题的方法:")
        print("1. 确保安装了所有必需的包")
        print("2. 检查输入文件是否存在并具有预期格式")
    
    logger.info("程序执行完毕")
    input("\n按Enter键返回主菜单...")

if __name__ == "__main__":
    # 设置基本日志配置
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    main()