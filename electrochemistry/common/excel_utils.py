"""
Excel处理工具模块
包含Excel文件创建、样式设置、数据写入等功能
"""
import logging
import subprocess
from typing import Tuple, Any, Dict
from datetime import datetime

logger = logging.getLogger(__name__)

# 检查openpyxl是否可用
try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def ensure_openpyxl_installed() -> None:
    """确保安装了openpyxl包"""
    global OPENPYXL_AVAILABLE
    
    if not OPENPYXL_AVAILABLE:
        try:
            # 确保导入成功
            import openpyxl
            from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
            OPENPYXL_AVAILABLE = True
        except ImportError:
            logger.error("缺少openpyxl包。正在安装...")
            try:
                subprocess.check_call(["pip", "install", "openpyxl"])
                # 安装后重新导入
                import openpyxl
                from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
                OPENPYXL_AVAILABLE = True
                logger.info("openpyxl安装成功")
            except Exception as e:
                logger.error(f"安装openpyxl失败: {str(e)}")
                raise

def setup_excel_workbook(sheet_name: str = "Data") -> Tuple[Any, Any, Any, Any, Any, Any]:
    """
    设置Excel工作簿并返回所需的样式
    
    参数:
        sheet_name: 工作表名称
    
    返回:
        (workbook, worksheet, header_fill, thin_border, center_aligned, openpyxl)
    """
    ensure_openpyxl_installed()
    
    # 确保openpyxl模块可用
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 设置列样式
    header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_aligned = Alignment(horizontal='center')
    
    return wb, ws, header_fill, thin_border, center_aligned, openpyxl

def get_excel_styles() -> Tuple[Any, Any, Any, Any]:
    """
    获取Excel样式，不创建新的工作簿
    
    返回:
        (header_fill, thin_border, center_aligned, openpyxl)
    """
    ensure_openpyxl_installed()
    
    # 确保openpyxl模块可用
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
    
    # 设置列样式
    header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_aligned = Alignment(horizontal='center')
    
    return header_fill, thin_border, center_aligned, openpyxl

def generate_output_filename(output_dir: str, basename: str, data_type: str) -> str:
    """
    生成带时间戳的输出文件名
    
    参数:
        output_dir: 输出目录
        basename: 基础文件名
        data_type: 数据类型 (cv, lsv, eis等)
    
    返回:
        完整的输出文件路径
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{output_dir}/{basename}_processed_{data_type}_data_{timestamp}.xlsx"

def apply_header_style(ws: Any, row: int, col: int, value: str, styles: Dict[str, Any]) -> None:
    """
    应用表头样式并设置值
    
    参数:
        ws: 工作表对象
        row: 行号
        col: 列号
        value: 单元格值
        styles: 样式字典，包含fill, border, alignment等
    """
    cell = ws.cell(row=row, column=col, value=value)
    if 'fill' in styles:
        cell.fill = styles['fill']
    if 'border' in styles:
        cell.border = styles['border']
    if 'alignment' in styles:
        cell.alignment = styles['alignment']
    if 'font' in styles:
        cell.font = styles['font']

# New additions for general styling utilities
# 新增通用样式工具
try:
    # import openpyxl # Not importing at top level to keep create_excel_file compatible if it doesn't use it
    # 不在顶层导入 openpyxl，以保持 create_excel_file 在不使用它时的兼容性
    from openpyxl.styles import Font
    # from openpyxl.utils import get_column_letter as default_get_column_letter # Renamed to avoid conflict if user defines it
    from openpyxl.utils import get_column_letter as default_get_column_letter # 重命名以避免用户定义时发生冲突
    OPENPYXL_STYLES_AVAILABLE = True
except ImportError:
    OPENPYXL_STYLES_AVAILABLE = False
    # Font = None # Define Font as None if import fails
    Font = None # 如果导入失败，则将 Font 定义为 None
    # default_get_column_letter = None # Define as None
    default_get_column_letter = None # 定义为 None

def get_bold_font():
    """
    Returns an openpyxl Font object configured for bold text.
    Returns None if openpyxl.styles.Font is not available.
    返回一个配置为粗体文本的 openpyxl Font 对象。
    如果 openpyxl.styles.Font 不可用，则返回 None。
    """
    if not OPENPYXL_STYLES_AVAILABLE or Font is None:
        # This path should ideally not be taken if openpyxl is a hard dependency and checked earlier.
        # 如果 openpyxl 是一个硬依赖项并且之前已检查过，则理想情况下不应采用此路径。
        # print("Warning: openpyxl.styles.Font not available for get_bold_font utility.")
        print("警告: openpyxl.styles.Font 对于 get_bold_font 工具不可用。")
        return None
    return Font(bold=True)

def style_cells_bold(cells_to_style):
    """
    Applies bold font to a list of cell objects.
    :param cells_to_style: A list or iterator of openpyxl cell objects.
    将粗体字体应用于单元格对象列表。
    :param cells_to_style: openpyxl 单元格对象的列表或迭代器。
    """
    if not OPENPYXL_STYLES_AVAILABLE:
        # print("Warning: openpyxl styles not available for style_cells_bold.")
        print("警告: openpyxl 样式对于 style_cells_bold 不可用。")
        return
    
    bold_font_instance = get_bold_font()
    if bold_font_instance:
        for cell in cells_to_style:
            # if cell: # Ensure cell is not None
            if cell: # 确保单元格不为 None
                cell.font = bold_font_instance

def style_row_bold(sheet, row_index, start_col=None, end_col=None):
    """
    Applies bold font to a specific row, optionally within a column range.
    :param sheet: openpyxl sheet object.
    :param row_index: 1-based row index.
    :param start_col: 1-based start column index (optional). If None, styles from the first column.
    :param end_col: 1-based end column index (optional). If None, styles up to sheet.max_column.
    将粗体字体应用于特定行，可选地在列范围内。
    :param sheet: openpyxl 工作表对象。
    :param row_index: 基于 1 的行索引。
    :param start_col: 基于 1 的起始列索引（可选）。如果为 None，则从第一列开始设置样式。
    :param end_col: 基于 1 的结束列索引（可选）。如果为 None，则样式设置到 sheet.max_column。
    """
    if not OPENPYXL_STYLES_AVAILABLE:
        # print("Warning: openpyxl styles not available for style_row_bold.")
        print("警告: openpyxl 样式对于 style_row_bold 不可用。")
        return

    bold_font_instance = get_bold_font()
    if not bold_font_instance:
        return

    if start_col is None:
        start_col = 1
    if end_col is None:
        end_col = sheet.max_column

    for col_idx in range(start_col, end_col + 1):
        cell = sheet.cell(row=row_index, column=col_idx)
        cell.font = bold_font_instance

def set_column_widths(sheet, col_width_map):
    """
    Sets column widths for a sheet.
    :param sheet: openpyxl sheet object.
    :param col_width_map: Dictionary of column letters to widths (e.g., {'A': 20, 'C': 15}).
    设置工作表的列宽。
    :param sheet: openpyxl 工作表对象。
    :param col_width_map: 列字母到宽度的字典（例如：{'A': 20, 'C': 15}）。
    """
    # if not OPENPYXL_STYLES_AVAILABLE: # Though column_dimensions doesn't strictly need Font
    if not OPENPYXL_STYLES_AVAILABLE: # 虽然 column_dimensions 并不严格需要 Font
        # print("Warning: openpyxl features might be limited for set_column_widths.")
        print("警告: openpyxl 功能可能对 set_column_widths 受限。")
        # return # Decide if to proceed or not; setting width might still work
        # return # 决定是否继续；设置宽度可能仍然有效
    
    for col_letter, width in col_width_map.items():
        try:
            sheet.column_dimensions[col_letter].width = width
        except Exception as e:
            # print(f"Error setting width for column {col_letter}: {e}")
            print(f"为列 {col_letter} 设置宽度时出错: {e}")