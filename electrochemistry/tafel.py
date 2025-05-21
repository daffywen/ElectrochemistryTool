# tafel.py
import logging
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Tuple, Optional

# Assuming excel_utils is in .common relative to this file's package
from .common import excel_utils

logger = logging.getLogger(__name__)

def _find_column_indices(ws: openpyxl.worksheet.worksheet.Worksheet, headers_to_find: List[str]) -> Dict[str, Optional[int]]:
    """Finds column indices for given headers in the first row of a worksheet."""
    indices: Dict[str, Optional[int]] = {header: None for header in headers_to_find}
    # Assuming headers are in the first row
    for col_idx, cell in enumerate(ws[1], 1): # ws[1] gets the first row
        if cell.value in indices:
            indices[cell.value] = col_idx
    
    for header, index in indices.items():
        if index is None:
            logger.warning(f"Header '{header}' not found in sheet '{ws.title}'.")
    return indices

def _read_column_data(ws: openpyxl.worksheet.worksheet.Worksheet, col_idx: Optional[int], start_row: int) -> np.ndarray:
    """Reads numeric data from a specified column starting from start_row."""
    data = []
    if col_idx is None:
        return np.array([])
        
    for r_idx in range(start_row, ws.max_row + 1):
        cell_value = ws.cell(row=r_idx, column=col_idx).value
        if cell_value is None: # Treat as end of data for this column or break
            # Depending on sheet structure, might be empty cells within data.
            # For now, let's assume contiguous data. If None, means end or sparse data.
            # If data can be sparse, this needs adjustment (e.g. append None and handle later)
            break 
        try:
            data.append(float(cell_value))
        except (ValueError, TypeError):
            logger.warning(f"Could not convert '{cell_value}' to float in sheet '{ws.title}', row {r_idx}, col {col_idx}. Skipping value.")
            # data.append(np.nan) # Optionally use NaN for unconvertible values
    return np.array(data)

def process_tafel_data(
    wb: openpyxl.Workbook,
    eis_analysis_results: List[Dict[str, Any]],
    processed_lsv_sheet_names: List[str],
    folder_basename: str
) -> None:
    """
    Processes LSV data to generate Tafel plots and writes them to a new 'Tafel Data' sheet
    with each dataset arranged side-by-side, separated by a blank column.
    """
    logger.info("--- Starting Tafel data processing ---")
    logger.info(f"Received {len(processed_lsv_sheet_names)} LSV sheet names for Tafel: {processed_lsv_sheet_names}")
    logger.info(f"Workbook sheets at start of Tafel processing: {list(wb.sheetnames)}")

    unique_rs_values: List[float] = []
    if eis_analysis_results:
        all_rs = [res.get('rs') for res in eis_analysis_results if res.get('rs') is not None]
        valid_rs = sorted(list(set(rs for rs in all_rs if isinstance(rs, (int, float)))))
        if valid_rs:
            unique_rs_values = valid_rs
            logger.info(f"Using {len(unique_rs_values)} unique Rs values for Tafel calculation: {unique_rs_values}")
        else:
            logger.warning("No valid numeric Rs values found in EIS results. Overpotential columns might use defaults or not be generated if applicable.")
    else:
        logger.warning("EIS analysis results are empty. Rs values not available. Overpotential columns will not be generated if they depend on Rs.")

    if not unique_rs_values: # This check might be redundant if the above already handles it, but ensures clarity
        logger.warning("No Rs values available from EIS. Tafel Overpotential columns requiring Rs will not be generated.")

    tafel_sheet_name = "Tafel Data"
    if tafel_sheet_name in wb.sheetnames:
        idx = wb.sheetnames.index(tafel_sheet_name)
        wb.remove(wb.worksheets[idx])
        logger.info(f"Removed existing sheet: '{tafel_sheet_name}'.")
    tafel_ws = wb.create_sheet(tafel_sheet_name)
    logger.info(f"Created new sheet: '{tafel_sheet_name}'.")
    
    header_fill, thin_border, center_aligned, styles_module = excel_utils.get_excel_styles()
    bold_font = excel_utils.get_bold_font()
    # left_aligned = styles_module.styles.Alignment(horizontal='left', vertical='center', wrap_text=True) # Not used in new layout
    right_aligned = styles_module.styles.Alignment(horizontal='right', vertical='center')

    potential_header_lsv = "Potential"
    current_density_header_lsv = "Current Density"

    # Layout definitions for Tafel sheet
    # title_row_num = 1 # No longer used as file_id is moved
    param_name_row_num = 1 # NEW: Parameter names on row 1
    param_unit_row_num = 2 # NEW: Parameter units on row 2
    param_specifics_row_num = 3 # NEW: Specific identifiers on row 3
    # Blank row will be param_specifics_row_num + 1
    data_start_row_num = 5 # NEW: Data starts on row 5 (1+1+1 for headers + 1 blank row)
    
    current_tafel_block_start_col = 1 # Starting column for the current dataset\\'s block in Tafel sheet
    max_data_rows_written_overall = 0 # Tracks max data rows for any dataset

    # Standard headers expected in LSV sheets (these must match what lsv.py produces in its first data row)
    potential_header_lsv = "Potential"
    current_density_header_lsv = "Current Density"

    # This outer loop is for processing multiple source sheets, though typically it will be just ["LSV Data"]
    for lsv_sheet_name in processed_lsv_sheet_names:
        logger.info(f"Processing LSV sheet: '{lsv_sheet_name}' for Tafel data.")
        if lsv_sheet_name not in wb.sheetnames:
            logger.warning(f"LSV data sheet '{lsv_sheet_name}' not found in workbook. Skipping for Tafel.")
            # Optionally write a message to Tafel sheet about this skip, if desired, at current_tafel_block_start_col
            continue
        
        lsv_ws = wb[lsv_sheet_name]
        logger.info(f"Starting scan of LSV sheet '{lsv_sheet_name}'. Max columns: {lsv_ws.max_column}, Max rows: {lsv_ws.max_row}")
        
        current_scan_col_lsv = 1 # For scanning columns in lsv_ws
        datasets_found_in_sheet = 0

        while current_scan_col_lsv <= lsv_ws.max_column:
            logger.info(f"Scanning LSV sheet at column: {current_scan_col_lsv}") # DETAILED LOG
            potential_cell_value = lsv_ws.cell(row=1, column=current_scan_col_lsv).value
            
            current_density_cell_value = None
            next_col_for_cd = current_scan_col_lsv + 1
            if next_col_for_cd <= lsv_ws.max_column:
                current_density_cell_value = lsv_ws.cell(row=1, column=next_col_for_cd).value
                logger.info(f"  Potential header at col {current_scan_col_lsv}: '{potential_cell_value}', Current Density header at col {next_col_for_cd}: '{current_density_cell_value}'") # DETAILED LOG
            else:
                logger.info(f"  Potential header at col {current_scan_col_lsv}: '{potential_cell_value}'. No next column for Current Density (max_column: {lsv_ws.max_column}).") # DETAILED LOG
                
            # Use str() and strip() for robustness in header comparison
            potential_val_str = str(potential_cell_value).strip() if potential_cell_value is not None else ""
            cd_val_str = str(current_density_cell_value).strip() if current_density_cell_value is not None else ""

            if potential_val_str == potential_header_lsv and cd_val_str == current_density_header_lsv:
                datasets_found_in_sheet += 1
                logger.info(f"Found LSV dataset pair #{datasets_found_in_sheet} at columns {current_scan_col_lsv} (\'{potential_header_lsv}\') and {next_col_for_cd} (\'{current_density_header_lsv}\') in \'{lsv_sheet_name}\'.")

                potential_col_idx_lsv = current_scan_col_lsv
                current_density_col_idx_lsv = next_col_for_cd
                
                file_id_to_use = f"{lsv_sheet_name} Dataset {datasets_found_in_sheet} (LSV Cols {get_column_letter(potential_col_idx_lsv)}-{get_column_letter(current_density_col_idx_lsv)})" # Fallback
                
                # Try to get file_id from row 3 of the current density column in LSV sheet
                file_id_cell_value = lsv_ws.cell(row=3, column=current_density_col_idx_lsv).value
                
                if file_id_cell_value is not None and str(file_id_cell_value).strip():
                    file_id_to_use = str(file_id_cell_value).strip()
                    logger.info(f"Read File ID \'{file_id_to_use}\' from \'{lsv_sheet_name}\' cell {get_column_letter(current_density_col_idx_lsv)}3.")
                else:
                    logger.warning(f"File ID in \'{lsv_sheet_name}\' cell {get_column_letter(current_density_col_idx_lsv)}3 is empty/None. Using fallback: \'{file_id_to_use}\'.")
                
                logger.info(f"Final File ID for this Tafel dataset: \'{file_id_to_use}\'.")

                potential_values = _read_column_data(lsv_ws, potential_col_idx_lsv, start_row=5)
                current_density_values_ma = _read_column_data(lsv_ws, current_density_col_idx_lsv, start_row=5)
                logger.info(f"Read {potential_values.size} potential values and {current_density_values_ma.size} current density values for '{file_id_to_use}'.")

                if not potential_values.size or not current_density_values_ma.size or potential_values.size != current_density_values_ma.size:
                    logger.error(f"Data reading error or length mismatch for '{file_id_to_use}' (Pot: {potential_values.size}, CD: {current_density_values_ma.size}). Skipping this dataset for Tafel.")
                    # Optionally write an error message to the Tafel sheet in the current block
                    # title_cell_text = f"Error: Data for {file_id_from_lsv_sheet}"
                    # error_cell = tafel_ws.cell(row=title_row_num, column=current_tafel_block_start_col, value=title_cell_text)
                    # if bold_font: error_cell.font = bold_font
                    # tafel_ws.merge_cells(start_row=title_row_num, 
                    #                          start_column=current_tafel_block_start_col, 
                    #                          end_row=title_row_num, 
                    #                          end_column=current_tafel_block_start_col) # Single cell merge
                    # current_tafel_block_start_col += 2 # Advance by 1 data col + 1 blank col
                    current_scan_col_lsv += 2 # Advance past this Pot/CD pair even if there's an error reading its data
                    continue
                
                current_density_values_a = current_density_values_ma / 1000.0
                log_j_values = np.log10(np.abs(current_density_values_a), where=np.abs(current_density_values_a)>0, out=np.full_like(current_density_values_a, np.nan))
                if np.all(np.isnan(log_j_values)):
                    logger.warning(f"All log(j) values are NaN for \'{file_id_to_use}\'. This might indicate all current densities were zero or invalid.")

                # Define parameter names, units, and specifics for Tafel output
                tafel_param_names = ["log"]
                tafel_param_units = ["j, A cm⁻²"]
                tafel_param_specifics = [""] # Blank for log column's 3rd row (specifics row)

                calculated_data_columns = [log_j_values] 

                if unique_rs_values:
                    for rs_val in unique_rs_values:
                        overpotential_v = np.abs((potential_values - 1.23) - (current_density_values_ma * rs_val * 0.001))
                        tafel_param_names.append("Overpotential")
                        tafel_param_units.append("V") # Unit for Overpotential
                        tafel_param_specifics.append(f"{file_id_to_use}, Rs={rs_val:.3f}") # Specifics for Overpotential
                        calculated_data_columns.append(overpotential_v)
                
                num_tafel_columns_for_this_dataset = len(tafel_param_names)

                # Write Parameter Names (Row 1)
                for header_idx, name_text in enumerate(tafel_param_names):
                    col_in_tafel = current_tafel_block_start_col + header_idx
                    cell = tafel_ws.cell(row=param_name_row_num, column=col_in_tafel, value=name_text)
                    if bold_font: cell.font = bold_font
                    if header_fill: cell.fill = header_fill
                    if thin_border: cell.border = thin_border
                    cell.alignment = center_aligned
                
                # Write Parameter Units (Row 2)
                for header_idx, unit_text in enumerate(tafel_param_units):
                    col_in_tafel = current_tafel_block_start_col + header_idx
                    cell = tafel_ws.cell(row=param_unit_row_num, column=col_in_tafel, value=unit_text)
                    if bold_font: cell.font = bold_font # Apply bold font to units
                    if header_fill: cell.fill = header_fill
                    if thin_border: cell.border = thin_border
                    cell.alignment = center_aligned

                # Write Parameter Specifics (Row 3)
                for header_idx, specific_text in enumerate(tafel_param_specifics):
                    col_in_tafel = current_tafel_block_start_col + header_idx
                    cell = tafel_ws.cell(row=param_specifics_row_num, column=col_in_tafel, value=specific_text)
                    # Style for specifics (can be bold if it contains file_id)
                    if bold_font and specific_text: cell.font = bold_font # Bold if not blank
                    if header_fill: cell.fill = header_fill
                    if thin_border: cell.border = thin_border
                    cell.alignment = center_aligned
                
                # Write One Blank Row with formatting (Row 4)
                for r_blank_offset in range(1): # For one blank row
                    blank_row_num = param_specifics_row_num + 1 + r_blank_offset # Row 4
                    for c_offset in range(num_tafel_columns_for_this_dataset):
                        col_in_tafel = current_tafel_block_start_col + c_offset
                        cell = tafel_ws.cell(row=blank_row_num, column=col_in_tafel, value=None) # Blank value
                        if header_fill: cell.fill = header_fill # Apply fill like headers
                        if thin_border: cell.border = thin_border # Apply border
                        cell.alignment = center_aligned # Apply alignment like headers
                
                # Write Data for this dataset in Tafel Sheet (Starts from data_start_row_num - Row 5)
                num_data_rows_this_dataset = len(calculated_data_columns[0]) if calculated_data_columns else 0
                for r_offset in range(num_data_rows_this_dataset):
                    for c_offset, data_array in enumerate(calculated_data_columns):
                        value_to_write = data_array[r_offset]
                        if isinstance(value_to_write, float) and np.isnan(value_to_write):
                            value_to_write = "NaN" 
                        col_in_tafel = current_tafel_block_start_col + c_offset
                        row_in_tafel = data_start_row_num + r_offset
                        cell = tafel_ws.cell(row=row_in_tafel, column=col_in_tafel, value=value_to_write)
                        if thin_border: cell.border = thin_border
                        cell.alignment = right_aligned
                        cell.number_format = '0.0000'
                
                max_data_rows_written_overall = max(max_data_rows_written_overall, num_data_rows_this_dataset)
                logger.info(f"Written {num_data_rows_this_dataset} rows of Tafel data for '{file_id_to_use}' starting at Tafel column {get_column_letter(current_tafel_block_start_col)}.")
                
                current_tafel_block_start_col += num_tafel_columns_for_this_dataset + 1 # +1 for blank separator column
                current_scan_col_lsv += 2 # Advance by 2 (Pot, CD) to find the next dataset in LSV sheet
            else:
                logger.info(f"  No dataset found starting at LSV sheet column {current_scan_col_lsv}. Potential='{potential_cell_value}' (Str: '{potential_val_str}'), CD='{current_density_cell_value}' (Str: '{cd_val_str}'). Advancing scan by 1.") # DETAILED LOG
                current_scan_col_lsv += 1
        
        logger.info(f"Finished scanning LSV sheet '{lsv_sheet_name}'. Found {datasets_found_in_sheet} dataset(s). Next Tafel block starts at column {get_column_letter(current_tafel_block_start_col)}")
        if datasets_found_in_sheet == 0:
            logger.warning(f"No datasets (matching '{potential_header_lsv}'/'{current_density_header_lsv}' headers in row 1) found in sheet '{lsv_sheet_name}'.")
            # Optionally write a message to Tafel sheet
            # no_data_msg_cell = tafel_ws.cell(row=title_row_num, column=current_tafel_block_start_col, value=f"No data from {lsv_sheet_name}")
            # current_tafel_block_start_col += 2 # Minimal advance if message written

    logger.info(f"Preparing to set column widths for '{tafel_sheet_name}'. Max data rows written: {max_data_rows_written_overall}. Total columns used up to: {get_column_letter(current_tafel_block_start_col-1) if current_tafel_block_start_col > 1 else 'None'}")
    
    if current_tafel_block_start_col > 1: # If any data was written
        col_widths = {}
        num_op_cols = len(unique_rs_values) # Number of overpotential columns per dataset
        columns_per_dataset_content = 1 + num_op_cols # log(j) + OPs
        
        # Iterate through the columns that were actually used or designated as separators
        for c_idx in range(1, current_tafel_block_start_col):
            # Determine if c_idx is a data column or a separator column
            # This logic assumes blocks are [data_cols] [separator_col] [data_cols] [separator_col] ...
            # Relative position within a "dataset_content + separator" group
            group_size_with_separator = columns_per_dataset_content + 1
            pos_in_group = (c_idx - 1) % group_size_with_separator # 0-indexed position

            if pos_in_group == 0: # First column of a dataset (log(j))
                col_widths[get_column_letter(c_idx)] = 20
            elif pos_in_group < columns_per_dataset_content: # Overpotential columns
                col_widths[get_column_letter(c_idx)] = 25
            elif pos_in_group == columns_per_dataset_content: # Blank separator column
                col_widths[get_column_letter(c_idx)] = 5
            # else case should not be reached if c_idx is correctly bounded by current_tafel_block_start_col

        if col_widths:
            logger.debug(f"Calculated col_widths map for '{tafel_sheet_name}': {col_widths}")
            logger.info(f"Attempting to call excel_utils.set_column_widths for sheet '{tafel_sheet_name}'...")
            excel_utils.set_column_widths(tafel_ws, col_widths)
            logger.info(f"Call to excel_utils.set_column_widths for sheet '{tafel_sheet_name}' completed.")
        else:
            logger.info(f"No column widths to set for '{tafel_sheet_name}'.")
    else:
        logger.info(f"Tafel Data sheet '{tafel_sheet_name}' is effectively empty; skipping column width setting.")

    logger.info("--- Tafel data processing completed ---")

