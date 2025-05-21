"""
循环伏安法(CV)数据处理模块
"""
import re
import os
import numpy as np
from typing import Tuple, List, Optional, Dict, Any
from pathlib import Path
import logging
from datetime import datetime
import sys

# 设置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# 导入共享模块
try:
    from .common import file_utils, excel_utils
except ImportError:
    # 如果作为独立模块运行
    try:
        # 添加父目录到路径
        parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        if parent_dir not in sys.path:
            sys.path.insert(0, parent_dir)
        from electrochemistry.common import file_utils, excel_utils
    except ImportError:
        print("无法导入共享模块，请确保目录结构正确")
        raise

# 可选的进度条支持
try:
    from tqdm import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False
    print("提示：安装tqdm包可以显示进度条。可以运行：pip install tqdm")

# 检测是否在 GUI 环境中运行（如 PyInstaller 打包后的应用）
def is_gui_mode():
    """判断是否在无控制台的GUI环境中运行"""
    # 检查是否通过 PyInstaller 打包
    if getattr(sys, 'frozen', False):
        return True
    # 检查是否有可用的控制台 
    try:
        # 尝试向标准输出写入，如果失败则可能在GUI模式中
        sys.stdout.write("")
        return False
    except (AttributeError, IOError):
        return True
    
# 设置日志
logger = logging.getLogger(__name__)

# 在GUI模式下禁用tqdm
if is_gui_mode():
    TQDM_AVAILABLE = False
    logger.info("检测到GUI模式，已禁用进度条显示")

# 导入openpyxl样式到模块级别
try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def is_cv_file(file_path: str) -> bool:
    """
    检查文件是否为循环伏安法(CV)数据文件
    通过查找文件内容中的关键词（如'Cyclic Voltammetry'）判断
    """
    try:
        with open(file_path, 'r', errors='ignore') as f:
            # # 只读取文件头部内容用于检测（提高性能）
            # 只读取文件头部内容用于检测（提高性能）
            header = f.read(1000)
            
            # # 明确匹配CV关键词
            # 明确匹配CV关键词
            is_cv = ('Cyclic Voltammetry' in header or 
                     'CYCLIC VOLTAMMETRY' in header)
            
            # # 排除非CV文件类型
            # 排除非CV文件类型
            is_not_cv = any(keyword in header for keyword in [
                'Linear Sweep Voltammetry', 
                'A.C. Impedance',
                'LSV',
                'Chronoamperometry',
                'Open Circuit',
                'EIS',
                'Tafel'
            ])
            
            # # 只有确认是CV且不包含非CV关键词的文件才返回True
            # 只有确认是CV且不包含非CV关键词的文件才返回True
            return is_cv and not is_not_cv
    except Exception as e:
        logger.warning(f"检查文件{file_path}时出错: {str(e)}")
        return False

def extract_scan_rate(content: str) -> Optional[int]:
    """从CV数据文件内容中提取扫描速率"""
    scan_rate_match = re.search(r'Scan Rate \(V/s\) = ([\d.]+)', content)
    if not scan_rate_match:
        # # 尝试其他可能的格式
        # 尝试其他可能的格式
        scan_rate_match = re.search(r'[Ss]can [Rr]ate\s*[:=]\s*([\d.]+)\s*[Vv]/s', content)
        if not scan_rate_match:
            return None
        
    scan_rate = float(scan_rate_match.group(1)) * 1000
    # return int(round(scan_rate))  # 转为整数，不保留小数
    return int(round(scan_rate))  # Convert to integer, no decimal places

def extract_potential_current_data(content: str) -> Tuple[List[float], List[float]]:
    """从CV数据文件内容中提取电位和电流数据"""
    # # 尝试多种可能的数据分隔形式
    # 尝试多种可能的数据分隔形式
    if 'Potential/V, Current/A' in content:
        data_section = content.split('Potential/V, Current/A')[1].strip()
    elif 'Potential' in content and 'Current' in content:
        # # 尝试找到数据部分的开始
        # 尝试找到数据部分的开始
        lines = content.strip().split('\n')
        data_start = -1
        for i, line in enumerate(lines):
            if ('Potential' in line and 'Current' in line) or \
               (re.search(r'[Pp]otential.*[Vv]', line) and re.search(r'[Cc]urrent.*[Aa]', line)):
                data_start = i + 1
                break
        
        if data_start > 0 and data_start < len(lines):
            data_section = '\n'.join(lines[data_start:])
        else:
            # # 如果找不到明确的标题行，尝试查找数字数据段
            # 如果找不到明确的标题行，尝试查找数字数据段
            potential_current_pattern = re.compile(r'(-?\d+\.?\d*)\s*,\s*(-?\d+\.?\d*)')
            matches = potential_current_pattern.findall(content)
            if matches:
                return [float(p) for p, c in matches], [float(c) for p, c in matches]
            return [], []
    else:
        return [], []
    
    lines = data_section.strip().split('\n')
    
    # # 预分配数组提高性能
    # 预分配数组提高性能
    potentials = []
    currents = []
    
    for line in lines:
        if line.strip():
            # # 尝试多种分隔符
            # 尝试多种分隔符
            if ',' in line:
                parts = line.strip().split(',')
            elif '\t' in line:
                parts = line.strip().split('\t')
            else:
                parts = line.strip().split()
            
            if len(parts) >= 2:
                try:
                    potential = float(parts[0].strip())
                    current = float(parts[1].strip())
                    potentials.append(potential)
                    currents.append(current)
                except (ValueError, IndexError):
                    continue
    
    return potentials, currents

def find_cycle_start_indices(potentials: List[float], tolerance: float = 1e-3) -> List[int]:
    """
    找到所有循环开始的索引
    现在能够自动检测循环的起始点，而不是硬编码特定电位值
    考虑到最后一个循环末尾可能有微小偏差(约0.001V)
    """
    # if not potentials or len(potentials) < 10:  # 至少需要一定数量的点
    if not potentials or len(potentials) < 10:  # At least a certain number of points are required
        return []
        
    cycle_starts = []
    
    # # 第一步：确定电位扫描的方向和范围
    # # 假设数据是从起始电位开始的（起始电位等于或接近结束电位）
    # 第一步：确定电位扫描的方向和范围
    # 假设数据是从起始电位开始的（起始电位等于或接近结束电位）
    start_potential = potentials[0]
    
    # # 找到极值点（最高和最低电位）
    # 找到极值点（最高和最低电位）
    min_potential = min(potentials)
    max_potential = max(potentials)
    
    # # 如果起始点接近最大或最小值，说明起始点不是循环中点
    # 如果起始点接近最大或最小值，说明起始点不是循环中点
    starts_at_extreme = (abs(start_potential - min_potential) < tolerance or 
                        abs(start_potential - max_potential) < tolerance)
    
    # # 确定电位变化的模式
    # 确定电位变化的模式
    direction_changes = []
    prev_direction = 0
    
    for i in range(1, len(potentials)):
        direction = 1 if potentials[i] > potentials[i-1] else (-1 if potentials[i] < potentials[i-1] else 0)
        if direction != 0 and direction != prev_direction and prev_direction != 0:
            # direction_changes.append(i-1)  # 记录方向改变的点
            direction_changes.append(i-1)  # Record the point where the direction changes
        if direction != 0:
            prev_direction = direction
    
    if not direction_changes:
        logger.warning("未能检测到电位扫描方向的变化，无法确定循环")
        return []
    
    # # 第二步：尝试找到循环的起始点
    # # 方法1：如果我们知道起始电位并且不是极值，找到接近起始电位的点
    # 第二步：尝试找到循环的起始点
    # 方法1：如果我们知道起始电位并且不是极值，找到接近起始电位的点
    if not starts_at_extreme:
        # for i in range(10, len(potentials)):  # 跳过前几个点
        for i in range(10, len(potentials)):  # Skip the first few points
            # # 检查是否回到接近起始电位的值（考虑可能的微小偏差）
            # 检查是否回到接近起始电位的值（考虑可能的微小偏差）
            if abs(potentials[i] - start_potential) < tolerance:
                cycle_starts.append(i)
    
    # # 方法2：如果没有找到足够的循环点，则使用方向变化点
    # 方法2：如果没有找到足够的循环点，则使用方向变化点
    if len(cycle_starts) <= 1 and direction_changes:
        # # 找出方向变化点的模式
        # 找出方向变化点的模式
        # if len(direction_changes) >= 4:  # 至少需要两个完整循环的方向变化点
        if len(direction_changes) >= 4:  # At least two complete cycles of direction change points are needed
            # # 计算相邻同类型方向变化点之间的间隔
            # 计算相邻同类型方向变化点之间的间隔
            intervals = []
            for i in range(2, len(direction_changes)):
                intervals.append(direction_changes[i] - direction_changes[i-2])
            
            if intervals:
                # # 使用平均间隔估计循环长度
                # 使用平均间隔估计循环长度
                avg_interval = sum(intervals) // len(intervals)
                
                # # 通过估计的循环长度找出可能的循环起点
                # 通过估计的循环长度找出可能的循环起点
                potential_after_first_extrema = potentials[direction_changes[0] + avg_interval//2]
                
                # # 找到所有接近这个特定电位的点
                # 找到所有接近这个特定电位的点
                for i in range(direction_changes[0], len(potentials), avg_interval//2):
                    if i + 10 < len(potentials) and abs(potentials[i] - potential_after_first_extrema) < tolerance:
                        cycle_starts.append(i)
    
    # # 方法3：如果前两种方法都没找到足够的循环点，使用极值点作为参考
    # 方法3：如果前两种方法都没找到足够的循环点，使用极值点作为参考
    if len(cycle_starts) <= 1:
        # # 找到所有接近最大值或最小值的点
        # 找到所有接近最大值或最小值的点
        reference = max_potential if abs(start_potential - max_potential) < abs(start_potential - min_potential) else min_potential
        points_near_ref = []
        
        for i in range(len(potentials)):
            if abs(potentials[i] - reference) < tolerance:
                points_near_ref.append(i)
        
        if len(points_near_ref) >= 2:
            # # 计算平均循环长度
            # 计算平均循环长度
            intervals = [points_near_ref[i+1] - points_near_ref[i] for i in range(len(points_near_ref)-1)]
            if intervals:
                avg_cycle = sum(intervals) // len(intervals)
                
                # # 以最小/最大值点为参考，回推可能的循环起点
                # 以最小/最大值点为参考，回推可能的循环起点
                for p in points_near_ref:
                    potential_start_idx = max(0, p - avg_cycle // 4)
                    cycle_starts.append(potential_start_idx)
    
    # # 最后，如果仍然没有找到循环起点，回退到原始的简单方法
    # 最后，如果仍然没有找到循环起点，回退到原始的简单方法
    if not cycle_starts:
        logger.warning("未能通过高级方法检测到循环起点，回退到简单检测")
        # # 检查是否数据分段明显（如每200个点一个循环）
        # 检查是否数据分段明显（如每200个点一个循环）
        distinct_segments = True
        segment_length = 0
        
        # # 尝试检测明显的数据分段
        # 尝试检测明显的数据分段
        for i in range(1, min(500, len(potentials))):
            if abs(potentials[i] - potentials[0]) < tolerance:
                segment_length = i
                break
        
        # if segment_length > 10:  # 合理的段长度
        if segment_length > 10:  # Reasonable segment length
            # # 检查后续点是否符合这个模式
            # 检查后续点是否符合这个模式
            for i in range(segment_length, len(potentials), segment_length):
                if i < len(potentials) and abs(potentials[i] - potentials[0]) >= tolerance:
                    distinct_segments = False
                    break
            
            if distinct_segments:
                cycle_starts = list(range(0, len(potentials), segment_length))
    
    # # 去除太靠近的点
    # 去除太靠近的点
    if cycle_starts:
        filtered_starts = [cycle_starts[0]]
        # min_distance = 10  # 最小间隔
        min_distance = 10  # Minimum interval
        
        for i in range(1, len(cycle_starts)):
            if cycle_starts[i] - filtered_starts[-1] >= min_distance:
                filtered_starts.append(cycle_starts[i])
        
        cycle_starts = filtered_starts
    
    logger.info(f"检测到的循环起点: {cycle_starts}")
    return cycle_starts

def extract_last_cycle(filename: str) -> Tuple[List[float], List[float], Optional[int]]:
    """提取CV数据文件中的最后一个循环数据"""
    try:
        with open(filename, 'r', errors='ignore') as f:
            content = f.read()
        
        # # 提取扫描速率
        # 提取扫描速率
        scan_rate = extract_scan_rate(content)
        
        # # 提取电位和电流数据
        # 提取电位和电流数据
        potentials, currents = extract_potential_current_data(content)
        
        if not potentials or not currents:
            raise ValueError(f"无法从文件{filename}中提取电位或电流数据")
        
        # # 找到循环开始点
        # 找到循环开始点
        cycle_starts = find_cycle_start_indices(potentials)
        
        logger.info(f"文件: {filename}, 总数据点: {len(potentials)}")
        logger.info(f"找到 {len(cycle_starts)} 个潜在循环起点，索引: {cycle_starts}")
        
        if len(cycle_starts) < 1:
            logger.warning(f"警告: 在{filename}中未找到循环起点，使用整个数据集")
            # # 如果找不到循环起点，使用整个数据集
            # 如果找不到循环起点，使用整个数据集
            processed_potentials = [p + 0.903 for p in potentials]
            processed_currents = [c * -1000 for c in currents]
            return processed_potentials, processed_currents, scan_rate
        
        # # 获取最后一个循环的起始索引
        # 获取最后一个循环的起始索引
        last_cycle_start_idx = cycle_starts[-1]
        
        # # 确定循环结束点（下一个循环的起始点或文件末尾）
        # 确定循环结束点（下一个循环的起始点或文件末尾）
        cycle_end_idx = len(potentials)
        
        # # 如果有明确的循环长度模式，使用该模式估计结束点
        # 如果有明确的循环长度模式，使用该模式估计结束点
        if len(cycle_starts) >= 2:
            avg_cycle_length = sum([cycle_starts[i] - cycle_starts[i-1] for i in range(1, len(cycle_starts))]) // (len(cycle_starts) - 1)
            estimated_end = last_cycle_start_idx + avg_cycle_length
            if estimated_end < len(potentials):
                cycle_end_idx = estimated_end
        
        # # 提取最后一个循环的所有数据
        # 提取最后一个循环的所有数据
        last_cycle_potentials = potentials[last_cycle_start_idx:cycle_end_idx]
        last_cycle_currents = currents[last_cycle_start_idx:cycle_end_idx]
        
        logger.info(f"已从{filename}中提取最后一个循环的{len(last_cycle_potentials)}个数据点")
        
        # # 应用转换：电位+0.903，电流*-1000
        # 应用转换：电位+0.903，电流*-1000
        processed_potentials = [p + 0.903 for p in last_cycle_potentials]
        processed_currents = [c * -1000 for c in last_cycle_currents]
        
        return processed_potentials, processed_currents, scan_rate
    
    except Exception as e:
        logger.error(f"处理文件{filename}时出错: {str(e)}")
        raise

def find_midpoint_values(potentials: List[float], currents: List[float], 
                         midpoint_value: float = 0.753, tolerance: float = 1e-3) -> Tuple[Optional[float], Optional[float]]:
    """
    找到电位为指定值(默认0.753V，对应-0.150V+0.903V)处的正向和反向扫描电流密度
    
    返回: (正向电流, 反向电流)
    """
    if not potentials or not currents:
        return None, None
    
    # # 自动确定中点电位：如果未找到指定电位，尝试使用数据范围的中点
    # 自动确定中点电位：如果未找到指定电位，尝试使用数据范围的中点
    target_found = False
    for p in potentials:
        if abs(p - midpoint_value) < tolerance:
            target_found = True
            break
    
    # # 如果未找到指定电位，计算数据范围的中点
    # 如果未找到指定电位，计算数据范围的中点
    if not target_found:
        min_potential = min(potentials)
        max_potential = max(potentials)
        auto_midpoint = (min_potential + max_potential) / 2
        logger.info(f"未在电位数据中找到{midpoint_value}V，自动使用中点电位{auto_midpoint:.3f}V")
        midpoint_value = auto_midpoint
        # # 增大容差以确保能找到点
        # 增大容差以确保能找到点
        tolerance = max(tolerance * 2, (max_potential - min_potential) / 50)
        
    # # 使用numpy向量化操作更高效
    # 使用numpy向量化操作更高效
    potentials_array = np.array(potentials)
    currents_array = np.array(currents)
    mask = np.abs(potentials_array - midpoint_value) < tolerance
    matches = np.where(mask)[0]
    
    if len(matches) < 2:
        logger.warning(f"找不到足够的{midpoint_value:.3f}V电位点（至少需要2个）")
        return None, None
    
    # # 分析电位扫描方向
    # 分析电位扫描方向
    directions = np.diff(potentials_array)
    direction_changes = np.where(np.diff(np.signbit(directions)))[0] + 1
    
    if len(direction_changes) < 1:
        # # 如果没有明显的方向变化，根据索引位置判断前后
        # 如果没有明显的方向变化，根据索引位置判断前后
        mid_idx = len(potentials) // 2
        first_half = [i for i in matches if i < mid_idx]
        second_half = [i for i in matches if i >= mid_idx]
        
        if first_half and second_half:
            forward_idx = first_half[0]
            reverse_idx = second_half[-1]
        else:
            # # 如果无法区分前后，就取第一个和最后一个匹配点
            # 如果无法区分前后，就取第一个和最后一个匹配点
            forward_idx = matches[0]
            reverse_idx = matches[-1]
    else:
        # # 根据电位变化方向判断
        # # 找到第一个方向变化点后的匹配点作为反向扫描点
        # 根据电位变化方向判断
        # 找到第一个方向变化点后的匹配点作为反向扫描点
        forward_matches = [i for i in matches if i < direction_changes[0]]
        reverse_matches = [i for i in matches if i >= direction_changes[0]]
        
        if not forward_matches or not reverse_matches:
            logger.warning("无法根据扫描方向区分正向和反向电流点")
            forward_idx = matches[0]
            reverse_idx = matches[-1]
        else:
            forward_idx = forward_matches[0]
            reverse_idx = reverse_matches[-1]
    
    forward_current = currents_array[forward_idx]
    reverse_current = currents_array[reverse_idx]
    
    logger.info(f"在电位{midpoint_value:.3f}V处找到：正向电流={forward_current:.6f}，反向电流={reverse_current:.6f}")
    return forward_current, reverse_current

def calculate_linear_fit(scan_rates: List[int], delta_j_values: List[float]) -> Tuple[float, float, float]:
    """
    计算扫描速率和Δj之间的线性关系
    
    返回: (斜率, 截距, R平方值)
    """
    # # 移除所有具有None值的配对
    # 移除所有具有None值的配对
    valid_pairs = [(rate, dj) for rate, dj in zip(scan_rates, delta_j_values) if dj is not None]
    
    if len(valid_pairs) < 2:
        return None, None, None
        
    x = np.array([pair[0] for pair in valid_pairs])
    y = np.array([pair[1] for pair in valid_pairs])
    
    # # 进行线性拟合: y = mx + b
    # 进行线性拟合: y = mx + b
    slope, intercept = np.polyfit(x, y, 1)
    
    # # 计算R平方值
    # 计算R平方值
    y_pred = slope * x + intercept
    ss_total = np.sum((y - np.mean(y)) ** 2)
    ss_residual = np.sum((y - y_pred) ** 2)
    r_squared = 1 - (ss_residual / ss_total)
    
    return slope, intercept, r_squared

def process_all_files_from_paths(file_paths: List[str], output_file: str, folder_basename: str, 
                             wb=None) -> Optional[Any]:
    """
    处理所有CV文件并保存为Excel格式，同时提取-0.150V处的数据点用于分析
    
    参数:
        file_paths: CV文件路径列表
        output_file: 输出Excel文件路径
        folder_basename: 文件夹基础名称
        wb: 可选，现有的工作簿对象（当与LSV模块集成时使用）
    
    返回:
        工作簿对象或None
    """
    # 提取并处理所有数据
    all_data = []
    scan_rates = []
    # file_names = []  # 仅用于内部处理，不在Excel中显示
    file_names = []  # Only for internal processing, not displayed in Excel
    # midpoint_data = []  # 将存储分析表的数据
    midpoint_data = []  # Will store data for the analysis table
    
    # # 记录传入的参数情况
    # 记录传入的参数情况
    logger.info(f"CV处理函数接收到的参数 - wb={wb is not None}")
    
    # # 使用进度条（如果可用）
    # 使用进度条（如果可用）
    file_iterator = tqdm(file_paths, desc="处理CV文件") if TQDM_AVAILABLE else file_paths
    
    for file_path in file_iterator:
        try:
            # # 获取不带扩展名的文件名(仅用于日志)
            # 获取不带扩展名的文件名(仅用于日志)
            full_file_name = os.path.basename(file_path)
            # file_name = os.path.splitext(full_file_name)[0]  # 去掉扩展名
            file_name = os.path.splitext(full_file_name)[0]  # Remove extension
            
            potentials, currents, scan_rate = extract_last_cycle(file_path)
            
            # # 跳过没有扫描速率的文件
            # 跳过没有扫描速率的文件
            if scan_rate is None:
                logger.warning(f"文件 {file_path} 中未找到扫描速率信息，跳过")
                continue
                
            all_data.append((potentials, currents))
            scan_rates.append(scan_rate)
            # file_names.append(file_name)  # 仅用于日志和内部引用
            file_names.append(file_name)  # For logging and internal reference only
            
            # # 找到电位为-0.150V的两个点（经过+0.903V校正后为0.753V）
            # 找到电位为-0.150V的两个点（经过+0.903V校正后为0.753V）
            forward_current, reverse_current = find_midpoint_values(potentials, currents)
            
            if forward_current is not None and reverse_current is not None:
                # # 计算delta_j为绝对差值，不除或乘以任何值
                # 计算delta_j为绝对差值，不除或乘以任何值
                delta_j = abs(reverse_current - forward_current)
                # # 不再包含文件名信息到midpoint_data
                # 不再包含文件名信息到midpoint_data
                midpoint_data.append((scan_rate, forward_current, reverse_current, delta_j))
                logger.info(f"文件 {file_path}: 扫描速率 {scan_rate} mV/s, 已找到电位点0.753V处的电流值")
            else:
                logger.warning(f"警告: 在{file_path}中找不到0.753V电位点")
                # # 找不到点时使用None，不包含文件名
                # 找不到点时使用None，不包含文件名
                midpoint_data.append((scan_rate, None, None, None))
            
        except Exception as e:
            logger.error(f"处理{file_path}时出错: {str(e)}")
            continue
    
    if not all_data:
        logger.error("没有找到任何可处理的CV数据文件")
        return None
    
    # # 按扫描速率排序
    # 按扫描速率排序
    combined_data = list(zip(all_data, scan_rates, midpoint_data))
    # combined_data.sort(key=lambda x: x[1])  # 按扫描速率排序
    combined_data.sort(key=lambda x: x[1])  # Sort by scan rate
    
    all_data = [item[0] for item in combined_data]
    scan_rates = [item[1] for item in combined_data]
    midpoint_data = [item[2] for item in combined_data]
    
    # 查找最大数据点数量
    max_length = max([len(potentials) for potentials, _ in all_data]) if all_data else 0
    
    # 使用提供的工作簿或创建新的
    # create_new_workbook_internally = False # No longer needed
    if wb is None:
        wb, ws, header_fill, thin_border, center_aligned, openpyxl_module = excel_utils.setup_excel_workbook("CV Data")
        # create_new_workbook_internally = True
        logger.info("创建新的工作簿用于CV数据")
    else:
        # 使用提供的工作簿但获取所需样式
        try:
            header_fill, thin_border, center_aligned, openpyxl_module = excel_utils.get_excel_styles()
        except (ImportError, AttributeError):
            # 如果找不到共享模块或函数，使用本地函数 (fallback, should ideally not happen)
            from openpyxl.styles import Alignment, Border, Side, PatternFill # Font removed, will use excel_utils.get_bold_font
            header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            center_aligned = Alignment(horizontal='center')
            openpyxl_module = openpyxl # Assuming openpyxl was imported at the top
        logger.info("使用现有工作簿处理CV数据")
    
    # Use the new utility for bold font
    bold_font = excel_utils.get_bold_font()

    # 检查工作表是否已存在或创建新的
    worksheet_name = "CV Data"
    if worksheet_name in wb.sheetnames:
        ws = wb[worksheet_name]
        logger.info(f"工作表 {worksheet_name} 已存在，将在此工作表上写入CV数据")
    else:
        ws = wb.create_sheet(worksheet_name)
        logger.info(f"已创建新的工作表: {worksheet_name} 用于CV数据")
    
    # 列从1开始
    current_col = 1
    
    # 第一列：Potential
    ws.cell(row=1, column=current_col).value = "Potential"
    ws.cell(row=2, column=current_col).value = "V vs. RHE"
    ws.cell(row=3, column=current_col).value = ""
    
    # 应用样式
    for row in range(1, 4):
        cell = ws.cell(row=row, column=current_col)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_aligned
        if bold_font: # Check if bold_font was successfully obtained
            cell.font = bold_font

    # 第四行：格式化的空行
    ws.cell(row=4, column=current_col).value = None
    ws.cell(row=4, column=current_col).fill = header_fill
    ws.cell(row=4, column=current_col).border = thin_border
    ws.cell(row=4, column=current_col).alignment = center_aligned

    # 其余列：不同扫描速率的Current Density
    for idx, scan_rate in enumerate(scan_rates):
        current_col_offset = idx + 1  # 因为第一列是Potential
        data_col = current_col + current_col_offset
        
        # 标题行
        ws.cell(row=1, column=data_col).value = "Current Density"
        ws.cell(row=2, column=data_col).value = "mA·cm⁻²"
        ws.cell(row=3, column=data_col).value = f"{scan_rate}mV/s"
        
        # 应用样式
        for row in range(1, 4):
            cell = ws.cell(row=row, column=data_col)
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_aligned
            if bold_font: # Check if bold_font was successfully obtained
                cell.font = bold_font

        # 第四行：格式化的空行
        ws.cell(row=4, column=data_col).value = None
        ws.cell(row=4, column=data_col).fill = header_fill
        ws.cell(row=4, column=data_col).border = thin_border
        ws.cell(row=4, column=data_col).alignment = center_aligned
    
    # 填充Potential列数据（所有扫描速率共用）
    if all_data:
        # 使用第一个数据集的电位数据（所有数据集理论上电位值相同）
        potentials = all_data[0][0]
        
        # 保存第一行数据用于后面复制到末尾
        first_row_potential = potentials[0] if potentials else None
        
        # 填充所有电位数据
        for row_idx, potential in enumerate(potentials, start=5):
            ws.cell(row=row_idx, column=current_col).value = potential
            ws.cell(row=row_idx, column=current_col).border = thin_border
        
        # 填充空白单元格（如果需要）
        for row_idx in range(5 + len(potentials), 5 + max_length):
            ws.cell(row=row_idx, column=current_col).value = None
            ws.cell(row=row_idx, column=current_col).border = thin_border
        
        # 在数据末尾添加第一行数据的两个副本，形成闭合循环
        if first_row_potential is not None:
            for i in range(2):
                last_row = 5 + max_length + i
                ws.cell(row=last_row, column=current_col).value = first_row_potential
                ws.cell(row=last_row, column=current_col).border = thin_border
    
    # 填充Current Density数据
    for idx, (potentials, currents) in enumerate(all_data):
        current_col_offset = idx + 1  # 因为第一列是Potential
        data_col = current_col + current_col_offset
        
        # 保存第一行数据用于后面复制到末尾
        first_row_current = currents[0] if currents else None
        
        # 填充所有电流数据
        for row_idx, current in enumerate(currents, start=5):
            ws.cell(row=row_idx, column=data_col).value = current
            ws.cell(row=row_idx, column=data_col).border = thin_border
        
        # 填充空白单元格（如果需要）
        for row_idx in range(5 + len(currents), 5 + max_length):
            ws.cell(row=row_idx, column=data_col).value = None
            ws.cell(row=row_idx, column=data_col).border = thin_border
        
        # 在数据末尾添加第一行数据的两个副本，形成闭合循环
        if first_row_current is not None:
            for i in range(2):
                last_row = 5 + max_length + i
                ws.cell(row=last_row, column=data_col).value = first_row_current
                ws.cell(row=last_row, column=data_col).border = thin_border
    
    # 设置更大的列宽
    col_letter = openpyxl_module.utils.get_column_letter(current_col)
    ws.column_dimensions[col_letter].width = 16  # Potential列
    for idx in range(len(scan_rates)):
        col_letter = openpyxl_module.utils.get_column_letter(current_col + idx + 1)
        ws.column_dimensions[col_letter].width = 16
    
    # 添加分析数据到同一工作表
    # 在数据列后跳过一列
    analysis_col = current_col + len(scan_rates) + 2
    
    # 为分析部分创建标题 - 移除File列
    ws.cell(row=1, column=analysis_col).value = "Scan rate"
    ws.cell(row=1, column=analysis_col+1).value = "Forward"
    ws.cell(row=1, column=analysis_col+2).value = "Reverse"
    ws.cell(row=1, column=analysis_col+3).value = "Δj"  # 修改标题

    ws.cell(row=2, column=analysis_col).value = "mV/s"
    ws.cell(row=2, column=analysis_col+1).value = "mA·cm⁻²"
    ws.cell(row=2, column=analysis_col+2).value = "mA·cm⁻²"
    ws.cell(row=2, column=analysis_col+3).value = "mA·cm⁻²"
    
    # 第三行仅有背景色
    for col in range(analysis_col, analysis_col+4):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.border = thin_border

    # 在Δj列第三行加上所选文件夹的名字
    ws.cell(row=3, column=analysis_col+3).value = folder_basename
    # 确保文件夹名称居中对齐
    ws.cell(row=3, column=analysis_col+3).alignment = center_aligned
    
    # 为分析标题应用样式
    for row in range(1, 3):  # 只有前两行有标题
        for col in range(analysis_col, analysis_col+4):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_aligned
            if bold_font: # Check if bold_font was successfully obtained
                cell.font = bold_font

    # 第四行：格式化的空行
    for col in range(analysis_col, analysis_col+4):
        cell = ws.cell(row=4, column=col)
        cell.value = None
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_aligned
    
    # 填充分析数据
    for idx, (scan_rate, forward, reverse, delta_j) in enumerate(midpoint_data, start=5):
        ws.cell(row=idx, column=analysis_col).value = scan_rate
        ws.cell(row=idx, column=analysis_col+1).value = forward
        ws.cell(row=idx, column=analysis_col+2).value = reverse
        ws.cell(row=idx, column=analysis_col+3).value = delta_j
        
        # 应用边框
        for col in range(analysis_col, analysis_col+4):
            ws.cell(row=idx, column=col).border = thin_border
    
    # 添加线性拟合数据
    # 提取扫描速率和delta_j数据，跳过None值
    valid_data = [(sr, dj) for sr, _, _, dj in midpoint_data if dj is not None]
    
    if len(valid_data) >= 2:  # 至少需要两个点才能进行线性拟合
        scan_rates_for_fit = [item[0] for item in valid_data]
        delta_j_for_fit = [item[1] for item in valid_data]
        
        slope, intercept, r_squared = calculate_linear_fit(scan_rates_for_fit, delta_j_for_fit)
        
        fit_row = 5 + len(midpoint_data) + 2  # 在数据下方留出空行
        
        ws.cell(row=fit_row, column=analysis_col).value = "Linear Fit Results:" # Changed to English
        ws.cell(row=fit_row+1, column=analysis_col).value = "Slope (m)" # Changed to English
        ws.cell(row=fit_row+1, column=analysis_col+1).value = slope
        
        ws.cell(row=fit_row+2, column=analysis_col).value = "Intercept (b)" # Changed to English
        ws.cell(row=fit_row+2, column=analysis_col+1).value = intercept
        
        ws.cell(row=fit_row+3, column=analysis_col).value = "R²"
        ws.cell(row=fit_row+3, column=analysis_col+1).value = r_squared
        
        # 计算公式 y = mx + b
        equation = f"y = {slope:.4f}x + {intercept:.4f} (R² = {r_squared:.4f})"
        ws.cell(row=fit_row+4, column=analysis_col).value = "Equation" # Changed to English
        ws.cell(row=fit_row+4, column=analysis_col+1).value = equation
        
        # 计算Cdl值 = 斜率 * 1000, 保留两位小数
        cdl_value = slope * 1000
        
        # 设置Cdl行的样式：加粗字体和黄色背景
        cdl_label_cell = ws.cell(row=fit_row+5, column=analysis_col, value="Cdl")
        cdl_value_cell = ws.cell(row=fit_row+5, column=analysis_col+1, value=f"{cdl_value:.2f} mF·cm⁻²")
        
        # 创建黄色填充 - openpyxl_module should be available from setup_excel_workbook or get_excel_styles
        yellow_fill = openpyxl_module.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # 应用样式到Cdl标签和值
        if bold_font: # Check if bold_font was successfully obtained
            cdl_label_cell.font = bold_font
            cdl_value_cell.font = bold_font
        cdl_label_cell.fill = yellow_fill
        cdl_value_cell.fill = yellow_fill
        
        # 添加边框保持一致性
        cdl_label_cell.border = thin_border
        cdl_value_cell.border = thin_border

        # Apply bold font to other linear fit labels
        for r_offset in range(5): # For "Linear Fit Results:", "Slope (m)", "Intercept (b)", "R²", "Equation"
            label_cell = ws.cell(row=fit_row + r_offset, column=analysis_col)
            if bold_font: # Check if bold_font was successfully obtained
                label_cell.font = bold_font
            label_cell.border = thin_border # Ensure border for labels
            # value_cell = ws.cell(row=fit_row + 1 + r_offset, column=analysis_col + 1) # Values start one row below for slope etc.
            # No, value cells are handled differently or not at all for the main title "Linear Fit Results:"
            if r_offset > 0: # Skip for "Linear Fit Results:" which has no direct value cell in this loop
                 value_cell_to_border = ws.cell(row=fit_row + r_offset, column=analysis_col + 1) # Corrected: value cells are at the same row as their labels for slope, intercept etc.
                 value_cell_to_border.border = thin_border # Ensure border for values

    else: # if slope is None or not enough data for fit
        cdl_value = None # Ensure cdl_value is None if not calculated
    
    # 设置分析列的列宽
    # for col_idx in range(analysis_col, analysis_col+4):
    #     col_letter = openpyxl_module.utils.get_column_letter(col_idx)
    #     ws.column_dimensions[col_letter].width = 16
    
    # Use the new utility for column widths
    col_width_map_data = {openpyxl_module.utils.get_column_letter(current_col): 16}
    for idx in range(len(scan_rates)):
        col_width_map_data[openpyxl_module.utils.get_column_letter(current_col + idx + 1)] = 16
    excel_utils.set_column_widths(ws, col_width_map_data)

    col_width_map_analysis = {}
    for col_idx in range(analysis_col, analysis_col+4):
        col_width_map_analysis[openpyxl_module.utils.get_column_letter(col_idx)] = 16
    excel_utils.set_column_widths(ws, col_width_map_analysis)
    
    # 返回工作簿和分析摘要
    cv_analysis_summary = {'cdl': cdl_value, 'folder_basename': folder_basename}
    return wb, cv_analysis_summary

def find_cv_files(folder_path: str) -> List[str]:
    """
    在指定文件夹中查找所有可能的CV数据文件
    现在通过检查文件内容判断是否为CV文件
    """
    cv_files = []
    
    # 获取文件夹中所有txt文件
    txt_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) 
                if f.lower().endswith('.txt')]
    
    if not txt_files:
        logger.warning(f"在{folder_path}中未找到任何.txt文件")
        return []
        
    logger.info(f"在{folder_path}中找到{len(txt_files)}个.txt文件，检查是否为CV数据...")
    
    # 使用进度条（如果可用）
    file_iterator = tqdm(txt_files, desc="识别CV数据文件") if TQDM_AVAILABLE else txt_files
    
    # 检查每个文件是否为CV数据文件
    for file_path in file_iterator:
        if is_cv_file(file_path):
            cv_files.append(file_path)
    
    return cv_files

def main():
    """主程序执行流程 (用于CV模块独立测试)"""
    # Setup basic logging for standalone run
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    logger.info("CV数据处理模块独立运行启动")
    
    # 提示用户选择文件夹
    folder_path = file_utils.select_folder() 
    
    if not folder_path:
        logger.warning("未选择文件夹，程序退出。")
        return
    
    logger.info(f"已选择文件夹: {folder_path}")
    
    # 在选定的文件夹中查找CV文件
    cv_files = find_cv_files(folder_path)
    
    if not cv_files:
        logger.error(f"在选定的文件夹中未找到任何CV数据文件。")
        print("\n在选定的文件夹中未找到任何CV数据文件。请确保文件内容包含循环伏安法相关信息。")
        return
    
    logger.info(f"找到 {len(cv_files)} 个CV数据文件:")
    for file_path in cv_files:
        logger.info(f"  {os.path.basename(file_path)}")
    
    # 输出文件路径
    folder_basename = os.path.basename(folder_path.rstrip("\\/"))
    # 新建 processed_data 文件夹（如果不存在）
    output_dir = os.path.join(folder_path, "processed_data")
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    # 添加时间戳到文件名以避免覆盖
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(output_dir, f"{folder_basename}_processed_cv_data_{timestamp}.xlsx")
    
    # 确保安装了openpyxl并设置全局变量
    excel_utils.ensure_openpyxl_installed()
    
    # 处理文件并保存结果
    try:
        processed_wb = process_all_files_from_paths(cv_files, output_file, folder_basename, None)
        if not processed_wb:
            # 如果我们遇到权限错误，尝试使用不同的文件名
            logger.warning(f"保存到 {output_file} 失败，尝试使用备用文件名。")
            alt_output_file = os.path.join(output_dir, f"{folder_basename}_processed_cv_data_new_{timestamp}.xlsx")
            logger.info(f"尝试备用文件名: {alt_output_file}")
            processed_wb = process_all_files_from_paths(cv_files, alt_output_file, folder_basename, None)
            
            if not processed_wb:
                print("\n建议解决问题的方法:")
                print("1. 关闭可能已打开的Excel文件")
                print("2. 检查文件夹的写入权限")
                print("3. 尝试以管理员权限运行脚本")
        else:
            print(f"\n处理完成！结果已保存到：{output_file}")
            print(f"共处理了 {len(cv_files)} 个CV数据文件")
    except Exception as e:
        logger.error(f"发生意外错误: {str(e)}")
        print(f"\n发生意外错误: {str(e)}")
        print("\n建议解决问题的方法:")
        print("1. 确保安装了所有必需的包")
        print("2. 检查输入文件是否存在并具有预期格式")
        print("3. 尝试以管理员权限运行脚本")
        
    logger.info("程序执行完毕")

# 主执行
if __name__ == "__main__":
    main()